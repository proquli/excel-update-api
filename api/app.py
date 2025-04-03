from flask import Flask, request, jsonify
import os
import json
import traceback
import openpyxl
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from google.oauth2 import service_account
import time
import signal
from functools import wraps
from threading import Thread

app = Flask(__name__)

# Set request timeout (in seconds)
REQUEST_TIMEOUT = 30

# Google Drive API scope
SCOPES = ['https://www.googleapis.com/auth/drive']

# Map of input fields to Excel cell references
EXCEL_CELL_MAP = {
    "projectName": "D29",  # Merged cells D29:G29
    "projectNumber": "D8",  # Merged cells D8:F8
    "branch": "D6"  # Merged cells D6:G6
}

def timeout_handler(signum, frame):
    raise TimeoutError("Request timed out")

def authenticate():
    """Authenticate with Google Drive API using service account."""
    try:
        # Get service account JSON from environment variable
        service_account_info = json.loads(os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON"))
        app.logger.info(f"Authenticating with service account: {service_account_info.get('client_email', 'unknown')}")
        
        credentials = service_account.Credentials.from_service_account_info(
            service_account_info, scopes=SCOPES)
        return credentials
    except Exception as e:
        app.logger.error(f"Authentication error: {str(e)}")
        app.logger.error(traceback.format_exc())
        raise

def parse_request_data(request):
    """Helper function to parse request data from various formats"""
    data = {}
    raw_data = request.get_data(as_text=True)
    
    # First check if we have form data
    if request.form:
        data = request.form.to_dict()
        app.logger.info(f"Parsed form data: {data}")
    # Then check for URL-encoded data with mismatched Content-Type
    elif "=" in raw_data and "&" in raw_data:
        # Parse URL-encoded data manually
        from urllib.parse import parse_qs
        parsed_data = parse_qs(raw_data)
        # Convert from lists to single values
        data = {k: v[0] for k, v in parsed_data.items()}
        app.logger.info(f"Parsed URL-encoded data: {data}")
    # Finally try JSON parsing
    elif raw_data:
        try:
            data = request.json
            app.logger.info(f"Parsed JSON data: {data}")
        except:
            app.logger.warning("Failed to parse as JSON")
            
    app.logger.info(f"Processed webhook data: {data}")
    return data


def download_excel(service, file_id, download_path):
    """Download Excel file from Google Drive with timeout handling."""
    try:
        # Add support for shared drives
        app.logger.info(f"About to download file with ID: '{file_id}'")
        
        try:
            file_metadata = service.files().get(
                fileId=file_id,
                supportsAllDrives=True
            ).execute()
            app.logger.info(f"File exists, name: {file_metadata.get('name')}")
        except Exception as e:
            app.logger.error(f"File metadata check failed: {str(e)}")
            raise ValueError(f"Unable to access file with ID {file_id}. Make sure the file exists and is shared with the service account.")
            
        request = service.files().get_media(
            fileId=file_id,
            supportsAllDrives=True
        )
        
        # Use context manager to ensure file is properly closed
        with open(download_path, 'wb') as fh:
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                status, done = downloader.next_chunk()
                app.logger.info(f"Download progress: {int(status.progress() * 100)}%")
                
        # Verify file is valid
        downloaded_size = os.path.getsize(download_path)
        app.logger.info(f"Downloaded file size: {downloaded_size} bytes")
        if downloaded_size < 5000:  # Adjust minimum expected size as needed
            raise ValueError(f"Downloaded file appears corrupt (too small): {downloaded_size} bytes")
            
        return True
    except Exception as e:
        app.logger.error(f"Download error: {str(e)}")
        app.logger.error(traceback.format_exc())
        raise

def update_excel(path, input_data):
    """Update specific cells in the Excel workbook based on input data."""
    try:
        wb = openpyxl.load_workbook(path, keep_vba=True, data_only=True)
        
        # Check if the required sheet exists
        if 'Project Setup Form' not in wb.sheetnames:
            raise Exception("Required sheet 'Project Setup Form' not found in the Excel file")
            
        sheet = wb['Project Setup Form']

        # Check if we have any mappable data
        updates_count = 0
        for field in EXCEL_CELL_MAP.keys():
            if field in input_data and input_data[field]:
                updates_count += 1
                
        if updates_count == 0:
            app.logger.warning("No mappable data found in input. Nothing to update.")
            return False

        # Create updates dictionary from input data
        updates = {}
        for field, cell_ref in EXCEL_CELL_MAP.items():
            if field in input_data and input_data[field]:
                updates[cell_ref] = input_data[field]
                app.logger.info(f"Updating {field} in cell {cell_ref} with value: {input_data[field]}")

        # Apply all updates
        for cell_ref, value in updates.items():
            sheet[cell_ref] = value

        # Save the workbook
        wb.save(path)
        
        # Explicitly close the workbook
        wb.close()
        
        # Verify file was saved properly
        try:
            verify_wb = openpyxl.load_workbook(path, keep_vba=True, data_only=True)
            verify_wb.close()
            app.logger.info(f"File verification successful: {path}")
        except Exception as e:
            app.logger.error(f"File verification failed: {str(e)}")
            raise ValueError(f"Excel file appears to be corrupted after save operation: {str(e)}")
        
        return True
        
    except Exception as e:
        app.logger.error(f"Excel update error: {str(e)}")
        app.logger.error(traceback.format_exc())
        raise

def upload_excel(service, file_id, path):
    """Upload updated Excel file back to Google Drive with timeout handling."""
    try:
        # Add support for shared drives
        file_metadata = service.files().get(
            fileId=file_id, 
            supportsAllDrives=True
        ).execute()
        app.logger.info(f"File metadata before upload: {file_metadata}")
        app.logger.info(f"File ID being used for API call: '{file_id}'")

        # Log file details
        file_size = os.path.getsize(path)
        app.logger.info(f"Local file size: {file_size} bytes")

        # Check if file appears valid
        if file_size < 5000:
            raise ValueError(f"File appears too small to be a valid Excel file: {file_size} bytes")

        # Use smaller chunk size for uploads
        media = MediaFileUpload(
            path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True,
            chunksize=512*1024  # Reduce to 512KB chunks
        )
        
        app.logger.info(f"Attempting chunked upload")
        app.logger.info(f"Attempting to upload file to ID: {file_id}")

        request = service.files().update(
            fileId=file_id,
            media_body=media,
            supportsAllDrives=True
        )
        
        response = None
        while response is None:
            status, response = request.next_chunk()
            if status:
                app.logger.info(f"Upload progress: {int(status.progress() * 100)}%")
        
        app.logger.info(f"Upload successful. Updated file details: {response}")
        return response
    except Exception as e:
        app.logger.error(f"Upload error: {str(e)}")
        app.logger.error(traceback.format_exc())
        raise

@app.route('/', methods=['GET', 'POST'])
def root():
    """Root endpoint that handles webhook requests from Zapier"""
    if request.method == 'POST':
        try:
            start_time = time.time()
            app.logger.info("Received webhook request")
            
            # Debug logging for request inspection
            raw_data = request.get_data(as_text=True)
            app.logger.info(f"Raw request data: {raw_data}")
            app.logger.info(f"Request form: {request.form}")
            app.logger.info(f"Request headers: {dict(request.headers)}")
            
            # Parse request data
            data = parse_request_data(request)
            
            # Process the webhook data
            if not data:
                return jsonify({"status": "error", "message": "No data provided"}), 400
                
            # Get the file ID
            file_id = data.get('Current File ID')
            if not file_id:
                return jsonify({"status": "error", "message": "No file ID provided"}), 400

            app.logger.info(f"Original file ID received: '{file_id}'")
            
            # Start processing in background
            Thread(target=process_excel_update, args=(file_id, data)).start()
            
            # Immediately return success to Zapier
            return jsonify({
                "status": "accepted", 
                "message": "Request accepted for processing"
            })
            
        except Exception as e:
            app.logger.error(f"Error processing webhook request: {str(e)}")
            app.logger.error(traceback.format_exc())
            return jsonify({"status": "error", "message": str(e)}), 500
    
    # Handle GET requests (like health checks)
    return jsonify({"status": "ok", "message": "API is running"}), 200

@app.route('/update-excel', methods=['POST'])
def update_excel_api():
    """API endpoint to update Excel file on Google Drive"""
    try:
        # Get the JSON data from the request
        data = request.json
        app.logger.info(f"Received request with data: {data}")
        
        # Validate the input
        if not data:
            return jsonify({"status": "error", "message": "No data provided"}), 400
            
        # Get the file ID
        file_id = data.get('Current File ID')
        if not file_id:
            return jsonify({"status": "error", "message": "No file ID provided"}), 400
            
        # Set up temporary file path
        temp_file = f"/tmp/{file_id}.xlsx"
        
        # Authenticate with Google Drive
        creds = authenticate()
        service = build('drive', 'v3', credentials=creds)
        
        # Download the file
        download_excel(service, file_id, temp_file)
        
        # Update the Excel file
        update_success = update_excel(temp_file, data)
        
        if update_success:
            # Upload the updated file
            upload_excel(service, file_id, temp_file)
            
        # Clean up
        if os.path.exists(temp_file):
            os.remove(temp_file)
            
        return jsonify({
            "status": "success",
            "message": "Excel file updated successfully" if update_success else "No updates were made"
        })
        
    except Exception as e:
        app.logger.error(f"Error processing request: {str(e)}")
        app.logger.error(traceback.format_exc())
        
        # Clean up in case of error
        if 'temp_file' in locals() and os.path.exists(temp_file):
            os.remove(temp_file)
            
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

@app.route('/test_file_access', methods=['GET'])
def test_file_access():
    file_id = request.args.get('file_id')
    app.logger.info(f"Testing file access for ID: '{file_id}'")
    try:
        creds = authenticate()
        service = build('drive', 'v3', credentials=creds)
        file_metadata = service.files().get(fileId=file_id, supportsAllDrives=True).execute()
        return jsonify({"success": True, "file_name": file_metadata.get('name')})
    except Exception as e:
        app.logger.error(f"Test file access error: {str(e)}")
        return jsonify({"success": False, "error": str(e)})

@app.route('/list_files', methods=['GET'])
def list_files():
    try:
        creds = authenticate()
        service = build('drive', 'v3', credentials=creds)
        results = service.files().list(
            pageSize=10,
            fields="files(id, name)",
            includeItemsFromAllDrives=True,
            supportsAllDrives=True
        ).execute()
        files = results.get('files', [])
        return jsonify({"files": files})
    except Exception as e:
        app.logger.error(f"List files error: {str(e)}")
        return jsonify({"error": str(e)})

@app.route('/test_connection', methods=['GET'])
def test_connection():
    try:
        creds = authenticate()
        service = build('drive', 'v3', credentials=creds)
        # Just list a few files to test connectivity
        results = service.files().list(pageSize=5).execute()
        return jsonify({"status": "success", "message": "Google Drive connection successful"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

def process_excel_update(file_id, data):
    """Background process to handle the Excel update"""
    try:
        app.logger.info(f"Starting background processing for file ID: {file_id}")
        start_time = time.time()
        
        temp_file = f"/tmp/{file_id}.xlsx"
        
        # Authenticate with Google Drive
        creds = authenticate()
        service = build('drive', 'v3', credentials=creds)
        
        # Download the file
        download_excel(service, file_id, temp_file)
        app.logger.info(f"Download completed in {time.time() - start_time:.2f} seconds")
        
        # Update the Excel file
        update_success = update_excel(temp_file, data)
        app.logger.info(f"Excel update completed in {time.time() - start_time:.2f} seconds")
        
        if update_success:
            # Upload the updated file
            upload_excel(service, file_id, temp_file)
            app.logger.info(f"Upload completed in {time.time() - start_time:.2f} seconds")
            
        # Clean up
        if os.path.exists(temp_file):
            os.remove(temp_file)
            
        end_time = time.time()
        app.logger.info(f"Background processing completed for file {file_id} in {end_time - start_time:.2f} seconds")
    except Exception as e:
        app.logger.error(f"Background processing error: {str(e)}")
        app.logger.error(traceback.format_exc())
        if 'temp_file' in locals() and os.path.exists(temp_file):
            os.remove(temp_file)

def verify_excel_file(file_path):
    with open(file_path, 'rb') as f:
        header = f.read(4)
    
    # Excel files start with these bytes (PKZip format)
    valid_header = header == b'PK\x03\x04'
    
    app.logger.info(f"Excel file header valid: {valid_header} for {file_path}")
    return valid_header

def diagnose_excel_file(file_path):
    """
    Diagnose potential issues with an Excel file
    Returns information about the file structure
    """
    import os
    import zipfile
    import xml.etree.ElementTree as ET
    
    results = {
        "file_exists": False,
        "file_size": 0,
        "is_valid_zip": False,
        "content_types": False,
        "workbook_xml": False,
        "worksheets": [],
        "custom_properties": False,
        "vba_content": False,
        "errors": []
    }
    
    try:
        # Basic file checks
        if not os.path.exists(file_path):
            results["errors"].append("File does not exist")
            return results
            
        results["file_exists"] = True
        results["file_size"] = os.path.getsize(file_path)
        
        if results["file_size"] < 2000:
            results["errors"].append(f"File too small: {results['file_size']} bytes")
        
        # Check if it's a valid ZIP file
        try:
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                results["is_valid_zip"] = True
                file_list = zip_ref.namelist()
                
                # Check for essential Office XML components
                if "[Content_Types].xml" in file_list:
                    results["content_types"] = True
                else:
                    results["errors"].append("Missing [Content_Types].xml")
                
                if "xl/workbook.xml" in file_list:
                    results["workbook_xml"] = True
                else:
                    results["errors"].append("Missing xl/workbook.xml")
                
                # Check for worksheets
                worksheets = [f for f in file_list if f.startswith("xl/worksheets/sheet")]
                results["worksheets"] = worksheets
                
                if not worksheets:
                    results["errors"].append("No worksheets found")
                
                # Check for VBA content
                vba_files = [f for f in file_list if "vbaProject" in f]
                results["vba_content"] = len(vba_files) > 0
                
                # Check for custom properties
                custom_props = [f for f in file_list if "customXml" in f]
                results["custom_properties"] = len(custom_props) > 0
                
        except zipfile.BadZipFile:
            results["is_valid_zip"] = False
            results["errors"].append("Not a valid ZIP file (Excel files are ZIP archives)")
            
        except Exception as e:
            results["errors"].append(f"Error examining ZIP structure: {str(e)}")
        
        # Final assessment
        if not results["errors"]:
            results["assessment"] = "File appears to be a valid Excel file"
        else:
            results["assessment"] = f"File has {len(results['errors'])} issues"
            
        return results
        
    except Exception as e:
        results["errors"].append(f"Diagnostic error: {str(e)}")
        return results

@app.route('/diagnose', methods=['GET'])
def diagnose_endpoint():
    """
    Endpoint to diagnose Excel file issues
    Usage: /diagnose?file_id=YOUR_FILE_ID
    """
    file_id = request.args.get('file_id')
    if not file_id:
        return jsonify({"error": "No file_id provided"}), 400
        
    try:
        # Set up temporary file path
        temp_file = f"/tmp/{file_id}_diagnostic.xlsx"
        
        # Authenticate with Google Drive
        creds = authenticate()
        service = build('drive', 'v3', credentials=creds)
        
        # Get file metadata
        file_metadata = service.files().get(
            fileId=file_id,
            fields='name,mimeType,size',
            supportsAllDrives=True
        ).execute()
        
        # Download the file
        try:
            download_excel(service, file_id, temp_file)
            download_success = True
        except Exception as e:
            download_success = False
            download_error = str(e)
        
        # Run diagnostics
        if download_success:
            diagnostic_results = diagnose_excel_file(temp_file)
            
            # Clean up
            if os.path.exists(temp_file):
                os.remove(temp_file)
                
            return jsonify({
                "file_metadata": file_metadata,
                "download_success": download_success,
                "diagnostic_results": diagnostic_results
            })
        else:
            return jsonify({
                "file_metadata": file_metadata,
                "download_success": download_success,
                "download_error": download_error
            }), 500
            
    except Exception as e:
        app.logger.error(f"Diagnostic error: {str(e)}")
        
        # Clean up in case of error
        if 'temp_file' in locals() and os.path.exists(temp_file):
            os.remove(temp_file)
            
        return jsonify({
            "error": str(e)
        }), 500

if __name__ == '__main__':
    # For local development only - use proper WSGI server in production
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))